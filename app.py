from flask import Flask, render_template, request
import pandas as pd
import pickle
from openpyxl import Workbook, load_workbook
from datetime import datetime

app = Flask(__name__)

# Load the trained model
with open('/Users/garvagarwal/Desktop/Deploy/svc_model.pkl', 'rb') as file:
    trained_model = pickle.load(file)

# Define mappings for categorical features
feature_mappings = {
    'Do you feel good about yourself as a person?': {'Yes': 1, 'No': 0},
    'Do you feel worthwhile?': {'Yes': 1, 'No': 0},
    'Do you feel you have a number of good qualities as a person?': {'Yes': 1, 'No': 0},
    'Had any thoughts of harming yourself?': {'Yes': 1, 'No': 0},
    'Have you felt depressed DURING your pregnancy?': {'Yes': 1, 'No': 0},
    'How mild or severe would you consider your depression was during pregnancy': {
        'Mild depression': 0,
        'Significant signs of depression': 1},
    'Have you talked to your provider about depression during pregnancy?': {'Yes': 1, 'No': 0},
    'Have you felt anxious during your pregnancy?': {'Yes': 1, 'No': 0},
    'How long did you feel anxious?': {
        'Not very often': 0,
        'Quite often': 1,
        'Most of the time': 2},
    'Have you ever been depressed BEFORE pregnancy?': {'Yes': 1, 'No': 0},
    'Have you been under a physician’s care for depression before being pregnant?': {'Yes': 1, 'No': 0},
    'Did the physician prescribe any medication for your depression before being pregnant?': {'Yes': 1, 'No': 0},
    'Was your pregnancy planned?': {'Yes': 1, 'No': 0},
    'Are you satisfied with your marriage/relationship?': {'Yes': 1, 'No': 0},
    'Are you currently experiencing any marital/ relationship problems?': {'Yes': 1, 'No': 0},
    'Do you feel you receive adequate emotional and instrumental (help with household chores and child care) support from your partner?': {'Yes': 1, 'No': 0},
    'Do you feel you can rely on your partner when you need help?': {'Yes': 1, 'No': 0},
    'Do you feel you can confide in your partner, family and friends?': {'Yes': 1, 'No': 0},
    'Is your infant experiencing any health problems?': {'Yes': 1, 'No': 0},
    'Are you having problems with your baby feeding?': {'Yes': 1, 'No': 0},
    'Are you having problems with your baby sleeping?': {'Yes': 1, 'No': 0},
    'Would you consider your baby irritable or fussy?': {'Yes': 1, 'No': 0},
    'Since your new baby was born, how often have you felt down, depressed, or hopeless?': {
        'Never': 0,
        'Sometimes': 1,
        'Quite often': 2,
        'Most of the time': 3},
    'Since your new baby was born, how often have you had little interest or little pleasure in doing things?': {
        'Never': 0,
        'Sometimes': 1,
        'Quite often': 2,
        'Most of the time': 3}
}

# Update the process_input_data function to handle all feature mappings
def process_input_data(input_data):
    # Initialize encoded_data dictionary
    encoded_data = {}
    
    # Process each key-value pair in input_data
    for key, value in input_data.items():
        if key in feature_mappings:
            encoded_data[key] = feature_mappings[key][value]
    
    # Create DataFrame from encoded input data
    input_df = pd.DataFrame([encoded_data])
    
    return input_df


# Define route for home page
@app.route('/')
def home():
    return render_template('index.html')

# Define route for prediction
@app.route('/predict', methods=['POST'])
def predict():
    # Get user input from form
    input_data = request.form.to_dict()
    
    # Process input data
    input_df = process_input_data(input_data)
    
    # Make predictions
    predictions = trained_model.predict(input_df)
    
    # Map predictions to labels
    depression_level_mapping = {
        0: 'No Depression',
        1: 'Mild Depression',
        2: 'Moderate Depression',
        3: 'Severe Depression'
    }
    
    predicted_label = depression_level_mapping[predictions[0]]
    
    # Append user input, questions, and predicted output to Excel file
    append_to_excel(input_data, predicted_label)
    
    return render_template('result.html', predicted_label=predicted_label)

def append_to_excel(input_data, predicted_label):
    # Load existing Excel file or create new if it doesn't exist
    try:
        workbook = load_workbook('predictions2.xlsx')
        sheet = workbook.active
    except FileNotFoundError:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(['Timestamp'] + list(input_data.keys()) + ['Predicted Depression Level'])
    
    # Prepare row data for appending
    row_data = [pd.datetime.now().strftime('%Y-%m-%d %H:%M:%S')] + list(input_data.values()) + [predicted_label]
    
    # Append row data to the Excel file
    sheet.append(row_data)
    
    # Save the updated Excel file
    workbook.save('predictions2.xlsx')

if __name__ == '__main__':
    app.run(debug=True)
